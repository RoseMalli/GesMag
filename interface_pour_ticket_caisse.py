from tkinter import *
import pandas as pd
from tkinter import messagebox
from openpyxl import *
from messages import *

def ticket_caisse():
    interface_ticket_de_caisse = Toplevel()
    interface_ticket_de_caisse.geometry('565x467')
    interface_ticket_de_caisse.title(string = 'GesMag')
    interface_ticket_de_caisse.configure(bg = 'white')

    x = 565
    y = 515

    fenetre_largeur = interface_ticket_de_caisse.winfo_screenwidth()
    fenetre_hauteur = interface_ticket_de_caisse.winfo_screenheight()

    largeur = (fenetre_largeur / 2) - (x / 2)
    hauteur = (fenetre_hauteur / 2) - (y /2)

    interface_ticket_de_caisse.geometry(f'{x}x{y}+{int(largeur)}+{int(hauteur)}')

    label_supprimer_caissier = Label(interface_ticket_de_caisse, text = "Ticket de caisse", font = ("Arial", 15, 'bold'), fg = 'white', bg = '#375D81', width = 19, height = 1, bd = 3, relief = 'groove')
    label_supprimer_caissier.grid(row = 0, columnspan = 2, pady = 10)

    frame = Frame(interface_ticket_de_caisse, width = 400, height = 2, bg = 'white')
    frame.grid(row = 1, columnspan = 2, sticky = E, padx = 35, pady = 10)

    label_date = Label(interface_ticket_de_caisse, text = "Date de vente :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_date.grid(row = 1, column = 0, sticky = W, padx = 35, pady = 10)

    zone_de_saisi_pour_jour = Entry(frame, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 6)
    zone_de_saisi_pour_jour.grid(row = 0, column = 0, sticky = W, padx = 4)
    zone_de_saisi_pour_jour.focus()
    label_pour_slash = Label(frame, text = "/", width = 1, font = ('bold'), bg = 'white')
    label_pour_slash.grid(row = 0, column = 1, padx = 1)
    zone_de_saisi_pour_mois = Entry(frame, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 6)
    zone_de_saisi_pour_mois.grid(row = 0, column = 2, sticky = E, padx = 4)
    label_pour_slash2 = Label(frame, text = "/",  width = 1, font = ('bold'), bg = 'white')
    label_pour_slash2.grid(row = 0, column = 3, padx = 1)
    zone_de_saisi_pour_annee = Entry(frame, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 7)
    zone_de_saisi_pour_annee.grid(row = 0, column = 4, sticky = E, padx = 4)

    label_identifiant_du_produit = Label(interface_ticket_de_caisse, text = "Identifiant du produit :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_identifiant_du_produit.grid(row = 2, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_identifiant_produit = Entry(interface_ticket_de_caisse, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_identifiant_produit.grid(row = 2, column = 1, sticky = E, padx = 35, pady = 10)

    label_libelle_du_produit = Label(interface_ticket_de_caisse, text = "Libellé du produit :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_libelle_du_produit.grid(row = 3, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_libelle_produit = Entry(interface_ticket_de_caisse, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_libelle_produit.grid(row = 3, column = 1, sticky = E, padx = 35, pady = 10)

    label_quantite_du_produit = Label(interface_ticket_de_caisse, text = "Quantité :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_quantite_du_produit.grid(row = 4, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_quantite_produit = Entry(interface_ticket_de_caisse, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_quantite_produit.grid(row = 4, column = 1, sticky = E, padx = 35, pady = 10)

    label_prix_unitaire_du_produit = Label(interface_ticket_de_caisse, text = "Prix unitaire :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_prix_unitaire_du_produit.grid(row = 5, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_prix_unitaire_produit = Entry(interface_ticket_de_caisse, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_prix_unitaire_produit.grid(row = 5, column = 1, sticky = E, padx = 35, pady = 10)

    label_prix_total = Label(interface_ticket_de_caisse, text = "Prix total :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_prix_total.grid(row = 6, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_prix_total = Entry(interface_ticket_de_caisse, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_prix_total.grid(row = 6, column = 1, sticky = E, padx = 35, pady = 10)

    label_qunatite_restante = Label(interface_ticket_de_caisse, text = "Quantite restante :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_qunatite_restante.grid(row = 7, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_quantite_restante = Entry(interface_ticket_de_caisse, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_quantite_restante.grid(row = 7, column = 1, sticky = E, padx = 35, pady = 10)

    zone_de_saisi_pour_libelle_produit.configure(state = DISABLED)
    zone_de_saisi_pour_prix_unitaire_produit.configure(state = DISABLED)
    zone_de_saisi_pour_prix_total.configure(state = DISABLED)
    zone_de_saisi_pour_quantite_restante.configure(state = DISABLED)

    label_montant_total = Label(interface_ticket_de_caisse, text = "Montant total :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_montant_total.grid(row = 8, column = 0, sticky = W, padx = 35, pady = 10)

    resultat = StringVar()
    label3 = Label(interface_ticket_de_caisse, text = "0", textvariable = resultat, bg = 'white')
    label3.grid(row = 8, column = 1, sticky = E, padx = 35, pady = 10)
    resultat.set(0)

    wb = load_workbook('stocks.xlsx')
    sheet = wb.active

    wb1 = load_workbook('ticket_de_caisse.xlsx')
    sheet1 = wb1.active


    def ajouter():
        verification()
        data = pd.read_excel('stocks.xlsx')
        trouver = data["Identifiant"].values
        if zone_de_saisi_pour_identifiant_produit.get() in trouver:
            if len(zone_de_saisi_pour_quantite_produit.get()) != 0:
                zone_de_saisi_pour_libelle_produit.configure(state = NORMAL)
                zone_de_saisi_pour_prix_unitaire_produit.configure(state = NORMAL)
                zone_de_saisi_pour_prix_total.configure(state = NORMAL)
                zone_de_saisi_pour_quantite_restante.configure(state = NORMAL)

                zone_de_saisi_pour_libelle_produit.delete(0, END)
                zone_de_saisi_pour_prix_unitaire_produit.delete(0, END)
                zone_de_saisi_pour_prix_total.delete(0, END)
                zone_de_saisi_pour_quantite_restante.delete(0, END)

                for cell in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, min_col = 1, max_col = 4, values_only = True):
                    if cell[0] == zone_de_saisi_pour_identifiant_produit.get():
                        zone_de_saisi_pour_libelle_produit.insert(0, cell[1])
                        zone_de_saisi_pour_prix_unitaire_produit.insert(0, cell[3])

                        if int(zone_de_saisi_pour_quantite_produit.get()) <= int(cell[2]):
                            resultat.set(round(float(resultat.get()) + (float(zone_de_saisi_pour_prix_unitaire_produit.get()) * int(zone_de_saisi_pour_quantite_produit.get())), 2))
                            prix_total = round(float(float(zone_de_saisi_pour_prix_unitaire_produit.get()) * int(zone_de_saisi_pour_quantite_produit.get())), 2)
                            prix_total = float(prix_total)
                            zone_de_saisi_pour_prix_total.insert(0, prix_total)
                            nouvelle_quantite = int(cell[2] - int(zone_de_saisi_pour_quantite_produit.get()))
                            nouvelle_quantite = int(nouvelle_quantite)
                            zone_de_saisi_pour_quantite_restante.insert(0, nouvelle_quantite)
                            for row in sheet:
                                remove(sheet, row)
                
                            current_row = sheet.max_row
                            sheet.cell(row = current_row + 1, column = 1).value = zone_de_saisi_pour_identifiant_produit.get()
                            sheet.cell(row = current_row + 1, column = 2).value = zone_de_saisi_pour_libelle_produit.get()
                            sheet.cell(row = current_row + 1, column = 3).value = nouvelle_quantite
                            sheet.cell(row = current_row + 1, column = 4).value = float(zone_de_saisi_pour_prix_unitaire_produit.get())
                                
                        elif int(cell[2]) == 0:
                            zone_de_saisi_pour_quantite_produit.delete(0, END)
                            zone_de_saisi_pour_quantite_produit.focus()
                            resultat.get()
                            zone_de_saisi_pour_libelle_produit.configure(state = DISABLED)
                            zone_de_saisi_pour_prix_unitaire_produit.configure(state = DISABLED)
                            zone_de_saisi_pour_prix_total.configure(state = DISABLED)
                            zone_de_saisi_pour_quantite_restante.configure(state = DISABLED)
                            pas_de_stocks()

                        elif int(zone_de_saisi_pour_quantite_produit.get()) > int(cell[2]):
                            zone_de_saisi_pour_quantite_produit.delete(0, END)
                            zone_de_saisi_pour_quantite_produit.focus()
                            resultat.get()
                            zone_de_saisi_pour_libelle_produit.configure(state = DISABLED)
                            zone_de_saisi_pour_prix_unitaire_produit.configure(state = DISABLED)
                            zone_de_saisi_pour_prix_total.configure(state = DISABLED)
                            zone_de_saisi_pour_quantite_restante.configure(state = DISABLED)
                            depasse_la_quantite_de_stock()
            
                zone_de_saisi_pour_libelle_produit.configure(state = DISABLED)
                zone_de_saisi_pour_prix_unitaire_produit.configure(state = DISABLED)
                zone_de_saisi_pour_prix_total.configure(state = DISABLED)
                zone_de_saisi_pour_quantite_restante.configure(state = DISABLED)

                wb.save('stocks.xlsx') 

            else:
                zone_de_saisi_pour_quantite_produit.focus()
                saisissez_une_quantite()

        elif len(zone_de_saisi_pour_identifiant_produit.get()) == 0:
            zone_de_saisi_pour_identifiant_produit.focus()
            saisissez_un_identifiant()

        else:
            zone_de_saisi_pour_identifiant_produit.delete(0, END)
            zone_de_saisi_pour_quantite_produit.delete(0, END)
            zone_de_saisi_pour_identifiant_produit.focus()
            identifiant_incorrect()
    
    def verification():
        caracteres_speciaux = [".", "+", "-", "*", "/", "!", "§", ":", ";", "?", ",", "&", "~", "#", '"', "'", "{", "(", "[", "|", "`", "_", "^", "@", ")", "°", "]", "=", "}", "$", "£", "¨", "µ", "%", "<", ">"] 

        if len(zone_de_saisi_pour_jour.get()) == 0:
            zone_de_saisi_pour_jour.focus()
            saisissez_une_date1()
        elif any(i.isupper() for i in zone_de_saisi_pour_jour.get()):
            zone_de_saisi_pour_jour.focus()
            pas_de_majuscles()
        elif any(i.islower() for i in zone_de_saisi_pour_jour.get()):
            zone_de_saisi_pour_jour.focus()
            pas_de_minuscles()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_jour.get()):
            zone_de_saisi_pour_jour.focus()
            pas_de_caracteres_speciaux()
        elif len(zone_de_saisi_pour_jour.get()) != 2:
            zone_de_saisi_pour_jour.focus()
            entre_1_et_31()
            
        elif len(zone_de_saisi_pour_mois.get()) == 0:
            zone_de_saisi_pour_mois.focus()
            saisissez_une_date1()
        elif any(i.isupper() for i in zone_de_saisi_pour_mois.get()):
            zone_de_saisi_pour_mois.focus()
            pas_de_majuscles()
        elif any(i.islower() for i in zone_de_saisi_pour_mois.get()):
            zone_de_saisi_pour_mois.focus()
            pas_de_minuscles()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_mois.get()):
            zone_de_saisi_pour_mois.focus()
            pas_de_caracteres_speciaux()
        elif len(zone_de_saisi_pour_mois.get()) != 2:
            zone_de_saisi_pour_mois.focus()
            entre_1_et_12()
        
        elif len(zone_de_saisi_pour_annee.get()) == 0:
            zone_de_saisi_pour_annee.focus()
            saisissez_une_date1()
        elif any(i.isupper() for i in zone_de_saisi_pour_annee.get()):
            zone_de_saisi_pour_annee.focus()
            pas_de_majuscles()
        elif any(i.islower() for i in zone_de_saisi_pour_annee.get()):
            zone_de_saisi_pour_annee.focus()
            pas_de_minuscles()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_annee.get()):
            zone_de_saisi_pour_annee.focus()
            pas_de_caracteres_speciaux()
        elif len(zone_de_saisi_pour_annee.get()) != 4:
            zone_de_saisi_pour_annee.focus()
            nombre_a_4()
    
    def valider():
        verification()
        
        current_row1 = sheet1.max_row
        sheet1.cell(row = current_row1 + 1, column = 1).value = zone_de_saisi_pour_jour.get() + "/" + zone_de_saisi_pour_mois.get() + "/" + zone_de_saisi_pour_annee.get()
        sheet1.cell(row = current_row1 + 1, column = 2).value = float(resultat.get())

        wb1.save('ticket_de_caisse.xlsx')

        zone_de_saisi_pour_libelle_produit.configure(state = NORMAL)
        zone_de_saisi_pour_prix_unitaire_produit.configure(state = NORMAL)
        zone_de_saisi_pour_prix_total.configure(state = NORMAL)
        zone_de_saisi_pour_quantite_restante.configure(state = NORMAL)


        zone_de_saisi_pour_jour.delete(0, END)
        zone_de_saisi_pour_mois.delete(0, END)
        zone_de_saisi_pour_annee.delete(0, END)
        zone_de_saisi_pour_identifiant_produit.delete(0, END)
        zone_de_saisi_pour_libelle_produit.delete(0, END)
        zone_de_saisi_pour_quantite_produit.delete(0, END)
        zone_de_saisi_pour_prix_unitaire_produit.delete(0, END)
        zone_de_saisi_pour_prix_total.delete(0, END)
        zone_de_saisi_pour_quantite_restante.delete(0, END)
        resultat.set(0)

        zone_de_saisi_pour_libelle_produit.configure(state = DISABLED)
        zone_de_saisi_pour_prix_unitaire_produit.configure(state = DISABLED)
        zone_de_saisi_pour_prix_total.configure(state = DISABLED)
        zone_de_saisi_pour_quantite_restante.configure(state = DISABLED)

        zone_de_saisi_pour_jour.focus()

        valide_avec_succes()

    
    def remove(sheet, row):
        for cell in row:
            if cell.value != zone_de_saisi_pour_identifiant_produit.get():
                return
            else:
                sheet.delete_rows(row[0].row, 1)

    def quitter():
        interface_ticket_de_caisse.destroy()

    bouton_pour_ajouter = Button(interface_ticket_de_caisse, text = "Ajouter", font = ('Arial', 12, 'bold'), width = 10, bd = 3, bg = 'white', fg = '#00BFFF', activebackground = '#00BFFF', activeforeground = 'white', cursor = 'hand2', command = ajouter)
    bouton_pour_ajouter.grid(row = 9, columnspan = 2, pady = 10)
    
    bouton_pour_valider = Button(interface_ticket_de_caisse, text = "Valider", font = ('Arial', 12, 'bold'), width = 10, bd = 3, bg = 'white', fg = '#6FC771', activebackground = '#6FC771', activeforeground = 'white', cursor = 'hand2', command = valider)
    bouton_pour_valider.grid(row = 9, column = 1, sticky = E, padx = 35, pady = 10)
    
    bouton_pour_quitter = Button(interface_ticket_de_caisse, text = "Quitter", font = ('Arial', 12, 'bold'), width = 10, bd = 3, bg = 'white', fg = '#F32665', activebackground = '#F32665', activeforeground = 'white', cursor = 'hand2', command = quitter)
    bouton_pour_quitter.grid(row = 9, column = 0, sticky = W, padx = 35, pady = 30)

    #interface_ticket_de_caisse.mainloop()
