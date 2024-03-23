from tkinter import *
from tkinter import messagebox
from openpyxl import *
import pandas as pd
from messages import *

def interface_ajout_managers():
    interface_pour_ajouter_un_manager = Tk()
    interface_pour_ajouter_un_manager.title("GesMag")
    interface_pour_ajouter_un_manager.configure(bg = 'white')


    x = 546
    y = 520


    fenetre_largeur = interface_pour_ajouter_un_manager.winfo_screenwidth()
    fenetre_hauteur = interface_pour_ajouter_un_manager.winfo_screenheight()


    largeur = (fenetre_largeur / 2) - (x / 2)
    hauteur = (fenetre_hauteur / 2) - (y /2)


    interface_pour_ajouter_un_manager.geometry(f'{x}x{y}+{int(largeur)}+{int(hauteur)}')


    label_ajouter_manager = Label(interface_pour_ajouter_un_manager, text = "Ajoutez un nouveau manager", font = ("Arial", 15, 'bold'), fg = 'white', bg = '#375D81', width = 30, height = 1, bd = 3, relief = 'groove')
    label_ajouter_manager.grid(row = 0, columnspan = 2, pady = 10)


    identifiant_et_login_manager = StringVar()
    label_pour_indentifiant_et_login_manager = Label(interface_pour_ajouter_un_manager, text = "Indentifiant :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_pour_indentifiant_et_login_manager.grid(row = 1, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_identifiant_et_login_manager = Entry(interface_pour_ajouter_un_manager, textvariable = identifiant_et_login_manager, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_identifiant_et_login_manager.grid(row = 1, column = 1, sticky = E, padx = 35, pady = 10)
    zone_de_saisi_pour_identifiant_et_login_manager.focus()



    label_pour_nom_manager = Label(interface_pour_ajouter_un_manager, text = "Nom :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_pour_nom_manager.grid(row = 2, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_nom_manager = Entry(interface_pour_ajouter_un_manager, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_nom_manager.grid(row = 2, column = 1, sticky = E, padx = 35, pady = 10)


    label_pour_prenom_manager = Label(interface_pour_ajouter_un_manager, text = "Prenom :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_pour_prenom_manager.grid(row = 3, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_prenom_manager = Entry(interface_pour_ajouter_un_manager, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_prenom_manager.grid(row = 3, column = 1, sticky = E, padx = 35, pady = 10)


    frame = Frame(interface_pour_ajouter_un_manager, width = 400, height = 2, bg = 'white')
    frame.grid(row = 4, columnspan = 2, sticky = E, padx = 35, pady = 10)


    label_pour_date_de_naissance = Label(interface_pour_ajouter_un_manager, text = "Date de naissance :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_pour_date_de_naissance.grid(row = 4, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_date_de_naissance_manager = Entry(frame, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 6)
    zone_de_saisi_pour_date_de_naissance_manager.grid(row = 0, column = 0, sticky = W, padx = 4)
    label_pour_slash = Label(frame, text = "/", width = 1, font = ('bold'), bg = 'white')
    label_pour_slash.grid(row = 0, column = 1, padx = 1)
    zone_de_saisi_pour_date_de_naissance_manager2 = Entry(frame, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 6)
    zone_de_saisi_pour_date_de_naissance_manager2.grid(row = 0, column = 2, sticky = E, padx = 4)
    label_pour_slash2 = Label(frame, text = "/",  width = 1, font = ('bold'), bg = 'white')
    label_pour_slash2.grid(row = 0, column = 3, padx = 1)
    zone_de_saisi_pour_date_de_naissance_manager3 = Entry(frame, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 7)
    zone_de_saisi_pour_date_de_naissance_manager3.grid(row = 0, column = 4, sticky = E, padx = 4)


    label_pour_adresse = Label(interface_pour_ajouter_un_manager, text = "Adresse :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_pour_adresse.grid(row = 5, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_adresse_manager = Entry(interface_pour_ajouter_un_manager, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_adresse_manager.grid(row = 5, column = 1, sticky = E, padx = 35, pady = 10)


    label_pour_code_postal_manager = Label(interface_pour_ajouter_un_manager, text = "Code postal :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_pour_code_postal_manager.grid(row = 6, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_code_postal_manager = Entry(interface_pour_ajouter_un_manager, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_code_postal_manager.grid(row = 6, column = 1, sticky = E, padx = 35, pady = 10)


    label_pour_login_manager = Label(interface_pour_ajouter_un_manager, text = "Login :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_pour_login_manager.grid(row = 7, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_login_manager = Entry(interface_pour_ajouter_un_manager, textvariable = identifiant_et_login_manager, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_login_manager.grid(row = 7, column = 1, sticky = E, padx = 35, pady = 10)


    label_pour_mot_de_passe_manager = Label(interface_pour_ajouter_un_manager, text = "Mot de passe :", font = ("Arial", 12, 'bold'), bg = 'white')
    label_pour_mot_de_passe_manager.grid(row = 8, column = 0, sticky = W, padx = 35, pady = 10)
    zone_de_saisi_pour_mot_de_passe_manager = Entry(interface_pour_ajouter_un_manager, font = ("Arial", 12), bg = '#e3e3e3', bd = 3, width = 27)
    zone_de_saisi_pour_mot_de_passe_manager.grid(row = 8, column = 1, sticky = E, padx = 35, pady = 10)


    wb = load_workbook('base_de_données_caissiers_et_managers.xlsx')
    sheet = wb.active
    data = pd.read_excel('base_de_données_caissiers_et_managers.xlsx')

    def ajouter():
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row = current_row + 1, column = 1).value = zone_de_saisi_pour_identifiant_et_login_manager.get()
        sheet.cell(row = current_row + 1, column = 2).value = zone_de_saisi_pour_nom_manager.get()
        sheet.cell(row = current_row + 1, column = 3).value = zone_de_saisi_pour_prenom_manager.get()
        sheet.cell(row = current_row + 1, column = 4).value = zone_de_saisi_pour_date_de_naissance_manager.get() + "/" + zone_de_saisi_pour_date_de_naissance_manager2.get() + "/" + zone_de_saisi_pour_date_de_naissance_manager3.get()
        sheet.cell(row = current_row + 1, column = 5).value = zone_de_saisi_pour_adresse_manager.get()
        sheet.cell(row = current_row + 1, column = 6).value = zone_de_saisi_pour_code_postal_manager.get()
        sheet.cell(row = current_row + 1, column = 7).value = zone_de_saisi_pour_login_manager.get()
        sheet.cell(row = current_row + 1, column = 8).value = zone_de_saisi_pour_mot_de_passe_manager.get()

        wb.save('base_de_données_caissiers_et_managers.xlsx')

        vider()
    

    def enregistrement_verification():
        caracteres_speciaux = [".", "+", "-", "*", "/", "!", "§", ":", ";", "?", ",", "&", "~", "#", '"', "'", "{", "(", "[", "|", "`", "_", "^", "@", ")", "°", "]", "=", "}", "$", "£", "¨", "µ", "%", "<", ">"] 


        if len(zone_de_saisi_pour_identifiant_et_login_manager.get()) == 0:
            zone_de_saisi_pour_identifiant_et_login_manager.focus()
            saisissez_un_identifiant()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_identifiant_et_login_manager.get()):
            zone_de_saisi_pour_identifiant_et_login_manager.focus()
            pas_de_caracteres_speciaux()
        elif any(i.islower() for i in zone_de_saisi_pour_identifiant_et_login_manager.get()):
            zone_de_saisi_pour_identifiant_et_login_manager.focus()
            pas_de_minuscles()
        elif not zone_de_saisi_pour_identifiant_et_login_manager.get().startswith('M'):
            zone_de_saisi_pour_identifiant_et_login_manager.focus()
            identifiant_par_m()
        elif not any(i.isdigit() for i in zone_de_saisi_pour_identifiant_et_login_manager.get()):
            zone_de_saisi_pour_identifiant_et_login_manager.focus()
            au_moins_un_chiffre()
        elif not (len(zone_de_saisi_pour_identifiant_et_login_manager.get()) > 5):
            zone_de_saisi_pour_identifiant_et_login_manager.focus()
            cinq_caracteres_mini()
        
        elif len(zone_de_saisi_pour_nom_manager.get()) == 0:
            zone_de_saisi_pour_nom_manager.focus()
            saisissez_un_nom()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_nom_manager.get()):
            zone_de_saisi_pour_nom_manager.focus()
            pas_de_caracteres_speciaux()
        elif any(i.isdigit() for i in zone_de_saisi_pour_nom_manager.get()):
            zone_de_saisi_pour_nom_manager.focus()
            pas_de_chiffres()
        elif any(i.islower() for i in zone_de_saisi_pour_nom_manager.get()):
            zone_de_saisi_pour_nom_manager.focus()
            nom_en_majuscule()


        elif len(zone_de_saisi_pour_prenom_manager.get()) == 0:
            zone_de_saisi_pour_prenom_manager.focus()
            saisissez_un_prenom()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_prenom_manager.get()):
            zone_de_saisi_pour_prenom_manager.focus()
            pas_de_caracteres_speciaux()
        elif any(i.isdigit() for i in zone_de_saisi_pour_prenom_manager.get()):
            zone_de_saisi_pour_prenom_manager.focus()
            pas_de_chiffres()
        elif not any(i.islower() for i in zone_de_saisi_pour_prenom_manager.get()):
            zone_de_saisi_pour_prenom_manager.focus()
            prenom_en_minuscle()
        elif not any(i.isupper() for i in zone_de_saisi_pour_prenom_manager.get()):
            zone_de_saisi_pour_prenom_manager.focus()
            prenom_commence_par_majuscle()
            


        elif len(zone_de_saisi_pour_date_de_naissance_manager.get()) == 0:
            zone_de_saisi_pour_date_de_naissance_manager.focus()
            saisissez_une_date()
        elif any(i.isupper() for i in zone_de_saisi_pour_date_de_naissance_manager.get()):
            zone_de_saisi_pour_date_de_naissance_manager.focus()
            pas_de_majuscles()
        elif any(i.islower() for i in zone_de_saisi_pour_date_de_naissance_manager.get()):
            zone_de_saisi_pour_date_de_naissance_manager.focus()
            pas_de_minuscles()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_date_de_naissance_manager.get()):
            zone_de_saisi_pour_date_de_naissance_manager.focus()
            pas_de_caracteres_speciaux()
        elif len(zone_de_saisi_pour_date_de_naissance_manager.get()) != 2:
            zone_de_saisi_pour_date_de_naissance_manager.focus()
            entre_1_et_31()


        elif len(zone_de_saisi_pour_date_de_naissance_manager2.get()) == 0:
            zone_de_saisi_pour_date_de_naissance_manager2.focus()
            saisissez_une_date()
        elif any(i.isupper() for i in zone_de_saisi_pour_date_de_naissance_manager2.get()):
            zone_de_saisi_pour_date_de_naissance_manager2.focus()
            pas_de_majuscles()
        elif any(i.islower() for i in zone_de_saisi_pour_date_de_naissance_manager2.get()):
            zone_de_saisi_pour_date_de_naissance_manager2.focus()
            pas_de_minuscles()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_date_de_naissance_manager2.get()):
            zone_de_saisi_pour_date_de_naissance_manager2.focus()
            pas_de_caracteres_speciaux()
        elif len(zone_de_saisi_pour_date_de_naissance_manager2.get()) != 2:
            zone_de_saisi_pour_date_de_naissance_manager2.focus()
            entre_1_et_12()


        elif len(zone_de_saisi_pour_date_de_naissance_manager3.get()) == 0:
            zone_de_saisi_pour_date_de_naissance_manager3.focus()
            saisissez_une_date()
        elif any(i.isupper() for i in zone_de_saisi_pour_date_de_naissance_manager3.get()):
            zone_de_saisi_pour_date_de_naissance_manager3.focus()
            pas_de_majuscles()
        elif any(i.islower() for i in zone_de_saisi_pour_date_de_naissance_manager3.get()):
            zone_de_saisi_pour_date_de_naissance_manager3.focus()
            pas_de_minuscles()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_date_de_naissance_manager3.get()):
            zone_de_saisi_pour_date_de_naissance_manager3.focus()
            pas_de_caracteres_speciaux()
        elif len(zone_de_saisi_pour_date_de_naissance_manager3.get()) != 4:
            zone_de_saisi_pour_date_de_naissance_manager3.focus()
            nombre_a_4()

        elif len(zone_de_saisi_pour_adresse_manager.get()) == 0:
            zone_de_saisi_pour_adresse_manager.focus()
            saisissez_une_adresse()

        elif len(zone_de_saisi_pour_code_postal_manager.get()) == 0:
            zone_de_saisi_pour_code_postal_manager.focus()
            saisissez_un_code_postal()
        elif any(i.isupper() for i in zone_de_saisi_pour_code_postal_manager.get()):
            zone_de_saisi_pour_code_postal_manager.focus()
            pas_de_majuscles()
        elif any(i.islower() for i in zone_de_saisi_pour_code_postal_manager.get()):
            zone_de_saisi_pour_code_postal_manager.focus()
            pas_de_minuscles()
        elif any(i in caracteres_speciaux for i in zone_de_saisi_pour_code_postal_manager.get()):
            zone_de_saisi_pour_code_postal_manager.focus()
            pas_de_caracteres_speciaux()
        elif len(zone_de_saisi_pour_code_postal_manager.get()) != 5:
            zone_de_saisi_pour_code_postal_manager.focus()
            nombre_a_5()


        elif len(zone_de_saisi_pour_mot_de_passe_manager.get()) == 0:
            zone_de_saisi_pour_mot_de_passe_manager.focus()
            saisissez_un_mot_de_passe()
        elif not any(i in caracteres_speciaux for i in zone_de_saisi_pour_mot_de_passe_manager.get()):
            zone_de_saisi_pour_mot_de_passe_manager.focus()
            au_moins_un_caractere_special()
        elif not any(i.islower() for i in zone_de_saisi_pour_mot_de_passe_manager.get()):
            zone_de_saisi_pour_mot_de_passe_manager.focus()
            au_moins_une_minuscule()
        elif not any(i.isupper() for i in zone_de_saisi_pour_mot_de_passe_manager.get()):
            zone_de_saisi_pour_mot_de_passe_manager.focus()
            au_moins_une_majuscule()
        elif not any(i.isdigit() for i in zone_de_saisi_pour_mot_de_passe_manager.get()):
            zone_de_saisi_pour_mot_de_passe_manager.focus()
            au_moins_un_chiffre()
        elif len(zone_de_saisi_pour_mot_de_passe_manager.get()) < 8:
            zone_de_saisi_pour_mot_de_passe_manager.focus()
            au_moins_8_caracteres()
        

        else:
            verification_dans_identifiant = data["Identifiant"].values
            verification_dans_mot_de_passe = data["Mot de passe"].values
            if zone_de_saisi_pour_identifiant_et_login_manager.get() in verification_dans_identifiant:
                zone_de_saisi_pour_identifiant_et_login_manager.focus() 
                identifiant_existe_deja()
            elif zone_de_saisi_pour_mot_de_passe_manager.get() in verification_dans_mot_de_passe:
                zone_de_saisi_pour_mot_de_passe_manager.focus() 
                mot_de_passe_existe_deja()
            else:
                ajouter()
                zone_de_saisi_pour_identifiant_et_login_manager.focus()
                enregistre_avec_succes()


    def vider():
        zone_de_saisi_pour_identifiant_et_login_manager.delete(0, END)
        zone_de_saisi_pour_nom_manager.delete(0, END)
        zone_de_saisi_pour_prenom_manager.delete(0, END)
        zone_de_saisi_pour_date_de_naissance_manager.delete(0, END)
        zone_de_saisi_pour_date_de_naissance_manager2.delete(0, END)
        zone_de_saisi_pour_date_de_naissance_manager3.delete(0, END)
        zone_de_saisi_pour_adresse_manager.delete(0, END)
        zone_de_saisi_pour_code_postal_manager.delete(0, END)
        zone_de_saisi_pour_login_manager.delete(0, END)
        zone_de_saisi_pour_mot_de_passe_manager.delete(0, END)
        zone_de_saisi_pour_identifiant_et_login_manager.focus()


    def quitter():
        interface_pour_ajouter_un_manager.destroy()


    bouton_pour_enregistrer = Button(interface_pour_ajouter_un_manager, text = "Enregistrer", font = ('Arial', 12, 'bold'), width = 10, bd = 3, bg = 'white', fg = '#6FC771', activebackground = '#6FC771', activeforeground = 'white', cursor = 'hand2', command = enregistrement_verification)
    bouton_pour_enregistrer.grid(row = 9, column = 1, sticky = E, padx = 35, pady = 30)


    bouton_pour_vider = Button(interface_pour_ajouter_un_manager, text = "Vider", font = ('Arial', 12, 'bold'), width = 10, bd = 3, bg = 'white', fg = '#F4D03F', activebackground = '#F4D03F', activeforeground = 'white', cursor = 'hand2', command = vider)
    bouton_pour_vider.grid(row = 9, columnspan = 2)


    bouton_pour_quitter = Button(interface_pour_ajouter_un_manager, text = "Quitter", font = ('Arial', 12, 'bold'), width = 10, bd = 3, bg = 'white', fg = '#F32665', activebackground = '#F32665', activeforeground = 'white', cursor = 'hand2', command = quitter)
    bouton_pour_quitter.grid(row = 9, column = 0, sticky = W, padx = 35, pady = 30)


#interface_pour_ajouter_un_manager.mainloop()